import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def clean_column_names(df):
    """
    Cleans the column names of a DataFrame by converting to lowercase, replacing spaces with
    underscores, ensuring only a single underscore between words, and removing miscellaneous symbols.

    :param df: The input DataFrame.
    :type df: pandas DataFrame

    :return: The DataFrame with cleaned column names.
    :rtype: pandas DataFrame
    """
    tmp = df.copy()

    # Clean and transform the column names
    cleaned_columns = (
        df.columns.str.lower()
        .str.normalize("NFD")
        .str.encode("ascii", "ignore")
        .str.decode("utf-8")
        .str.replace(
            r"[^\w\s]", "", regex=True
        )  # Remove all non-alphanumeric characters except spaces
        .str.replace(r"\s+", "_", regex=True)  # Replace spaces with underscores
        .str.replace(r"_+", "_", regex=True)  # Ensure only a single underscore between words
        .str.strip("_")
    )

    # Assign the cleaned column names back to the DataFrame
    tmp.columns = cleaned_columns

    return tmp


def limpiar_columna_texto(serie):
    return (
        serie.str.upper()
        .str.strip()
        .str.normalize("NFD")
        .str.encode("ascii", "ignore")
        .str.decode("utf-8")
    )


def guardar_dataframe_como_tabla_excel(
    df: pd.DataFrame,
    ruta_archivo: str,
    nombre_tabla: str = "Tabla1",
    estilo_tabla: str = "TableStyleMedium9",
) -> None:
    """
    Guarda un DataFrame de pandas en un archivo Excel como una tabla formal (no solo como
    rango de celdas).

    Parámetros:
        df (pd.DataFrame): DataFrame que quieres exportar.
        ruta_archivo (str): Ruta donde se guardará el archivo Excel.
        nombre_tabla (str): Nombre de la tabla en Excel (debe ser único y sin espacios).
        estilo_tabla (str): Estilo visual de la tabla en Excel. Por defecto "TableStyleMedium9".
    """
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active

    # Escribir el DataFrame en la hoja
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Calcular el rango de la tabla
    max_col = get_column_letter(df.shape[1])
    max_row = df.shape[0] + 1  # +1 para la cabecera
    rango_tabla = f"A1:{max_col}{max_row}"

    # Crear la tabla
    tabla = Table(displayName=nombre_tabla, ref=rango_tabla)

    # Aplicar estilo
    estilo = TableStyleInfo(
        name=estilo_tabla,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabla.tableStyleInfo = estilo

    # Agregar la tabla a la hoja
    ws.add_table(tabla)

    # Guardar el archivo
    wb.save(ruta_archivo)

    print(f"✅ Archivo guardado exitosamente en: {ruta_archivo}")


def asignar_correlativos(
    df, correlativo_inicial, glosas_con_correlativos, columna_salida="correlativo_2025"
):
    """
    Asigna correlativos secuenciales a los registros cuyo valor en la columna 'propiedad'
    esté presente en `glosas_con_correlativos`.

    Parámetros:
    - df: DataFrame con los datos.
    - correlativo_inicial: número entero desde el cual se empezarán a asignar los correlativos.
    - glosas_con_correlativos: lista de valores que indican las filas a las que se les debe asignar
    correlativo.
    - columna_salida: nombre de la columna donde se asignarán los correlativos. Por defecto:
    'correlativo_2025'.

    Retorna:
    - df_copy: copia del DataFrame con la columna de correlativos asignada.
    - ultimo_correlativo: último correlativo numérico asignado.
    """
    df_copy = df.copy()

    # Identifica las filas a las que se asignará correlativo
    mask = df_copy["propiedad"].isin(glosas_con_correlativos)
    df_filtrado = df_copy[mask]

    # Cantidad de registros a asignar
    cantidad = len(df_filtrado)

    # Genera correlativos
    correlativos = [f"2025-{i}" for i in range(correlativo_inicial, correlativo_inicial + cantidad)]

    # Asigna correlativos
    df_copy.loc[mask, columna_salida] = correlativos

    # Reemplaza "nan" como string por np.nan (por si acaso)
    df_copy[columna_salida] = df_copy[columna_salida].replace("nan", np.nan)

    # Calcula el último correlativo asignado
    ultimo_correlativo = (
        correlativo_inicial + cantidad - 1 if cantidad > 0 else correlativo_inicial - 1
    )

    return df_copy, ultimo_correlativo
